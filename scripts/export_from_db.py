#!/usr/bin/env python3
"""Тестовый скрипт: экспорт данных из существующей БД в Excel.

Работает полностью автономно — не требует запуска парсера.
Подключается к SQLite-базе, читает все объявления и формирует
Excel-файл с точно такой же структурой, как основной pipeline.

Использование:
    python scripts/export_from_db.py

Опционально можно указать пути:
    python scripts/export_from_db.py --db data/avito_listings.db --out data/avito_report.xlsx
"""

import argparse
import json
import sqlite3
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("Ошибка: openpyxl не установлен.")
    print("Установите: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────
# Модели (копия из src/models/product.py)
# ──────────────────────────────────────────────

class RoomCategory(Enum):
    ROOM = "Комната"
    STUDIO = "Студия"
    ONE = "1к"
    TWO = "2к"
    THREE = "3к"
    FOUR_PLUS = "4к+"
    UNKNOWN = "Неизвестно"


@dataclass
class RawListing:
    external_id: str
    latitude: float
    longitude: float
    room_category: RoomCategory
    price_60_days: list[int]
    calendar_60_days: list[int]
    snapshot_timestamp: datetime = field(
        default_factory=lambda: datetime.now(timezone.utc)
    )
    last_host_update: datetime | None = None
    min_stay: int = 1
    is_instant_book: bool = False
    host_rating: float = 0.0
    url: str = ""
    title: str = ""

    @property
    def full_url(self) -> str:
        base = "https://www.avito.ru"
        if self.url.startswith("http"):
            return self.url
        return f"{base}{self.url}"

    @property
    def occupancy_rate(self) -> float:
        if not self.calendar_60_days:
            return 0.0
        occupied = sum(self.calendar_60_days)
        return occupied / len(self.calendar_60_days)

    @property
    def average_price(self) -> float:
        if not self.price_60_days or not self.calendar_60_days:
            return 0.0
        prices = [
            price
            for price, occupied in zip(self.price_60_days, self.calendar_60_days)
            if price > 0 and occupied == 0
        ]
        if not prices:
            return 0.0
        return sum(prices) / len(prices)


# ──────────────────────────────────────────────
# Чтение из SQLite
# ──────────────────────────────────────────────

def deserialize_datetime(value: str | None) -> datetime | None:
    if value is None:
        return None
    dt = datetime.fromisoformat(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def deserialize_json_list(value: str | None) -> list[int]:
    if value is None:
        return []
    try:
        result = json.loads(value)
        if isinstance(result, list):
            return [int(x) for x in result]
        return []
    except (json.JSONDecodeError, ValueError):
        return []


def row_to_listing(row: sqlite3.Row) -> RawListing:
    snapshot_dt = deserialize_datetime(row["snapshot_timestamp"])
    if snapshot_dt is None:
        snapshot_dt = datetime.now(timezone.utc)

    # Безопасное преобразование room_category
    try:
        room_cat = RoomCategory(row["room_category"])
    except ValueError:
        room_cat = RoomCategory.UNKNOWN

    return RawListing(
        external_id=row["external_id"],
        latitude=float(row["latitude"]),
        longitude=float(row["longitude"]),
        room_category=room_cat,
        price_60_days=deserialize_json_list(row["price_60_days"]),
        calendar_60_days=deserialize_json_list(row["calendar_60_days"]),
        snapshot_timestamp=snapshot_dt,
        last_host_update=deserialize_datetime(row["last_host_update"]),
        min_stay=int(row["min_stay"]),
        is_instant_book=bool(row["is_instant_book"]),
        host_rating=float(row["host_rating"]),
        url=row["url"] or "",
        title=row["title"] or "",
    )


def load_all_listings(db_path: str) -> list[RawListing]:
    """Читает все объявления из SQLite и возвращает список RawListing."""
    if not Path(db_path).exists():
        print(f"Ошибка: файл базы данных не найден: {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    try:
        cursor = conn.execute(
            "SELECT * FROM listings ORDER BY snapshot_timestamp DESC"
        )
        rows = cursor.fetchall()
        listings = [row_to_listing(row) for row in rows]
        print(f"Загружено объявлений из БД: {len(listings)}")
        return listings
    except sqlite3.Error as e:
        print(f"Ошибка чтения из БД: {e}")
        sys.exit(1)
    finally:
        conn.close()


# ──────────────────────────────────────────────
# Экспорт в Excel (копия логики ExportService)
# ──────────────────────────────────────────────

REPORT_COLUMNS: list[tuple[str, int]] = [
    ("ID объявления", 18),
    ("Название", 45),
    ("Категория жилья", 16),
    ("Широта", 12),
    ("Долгота", 12),
    ("Средняя цена (руб./сут.)", 22),
    ("Занятость (%)", 14),
    ("Мин. срок (сут.)", 16),
    ("Мгновенное бронирование", 22),
    ("Рейтинг хоста", 14),
    ("Последнее обновление хостом", 26),
    ("Цены 60 дней", 80),
    ("Календарь 60 дней", 80),
    ("Ссылка", 55),
    ("Дата снимка", 20),
]


def format_array_semicolon(values: list[int]) -> str:
    if not values:
        return "—"
    return ";".join(str(v) for v in values)


def listing_to_row(listing: RawListing) -> list[str | int | float]:
    snapshot_date = listing.snapshot_timestamp.strftime("%Y-%m-%d %H:%M")

    last_update = ""
    if listing.last_host_update is not None:
        last_update = listing.last_host_update.strftime("%Y-%m-%d %H:%M")

    prices_str = format_array_semicolon(listing.price_60_days)
    calendar_str = format_array_semicolon(listing.calendar_60_days)

    return [
        listing.external_id,
        listing.title,
        listing.room_category.value,
        listing.latitude,
        listing.longitude,
        round(listing.average_price),
        listing.occupancy_rate,
        listing.min_stay,
        "Да" if listing.is_instant_book else "Нет",
        listing.host_rating,
        last_update,
        prices_str,
        calendar_str,
        listing.full_url,
        snapshot_date,
    ]


def write_header(ws: Worksheet) -> None:
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_index, (title, _) in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def write_data(ws: Worksheet, listings: list[RawListing]) -> None:
    data_alignment = Alignment(vertical="top", wrap_text=False)
    number_alignment = Alignment(horizontal="right", vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    link_font = Font(
        name="Calibri", size=10, color="0563C1", underline="single"
    )
    data_font = Font(name="Calibri", size=10)

    for row_index, listing in enumerate(listings, start=2):
        row_data = listing_to_row(listing)

        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        # Столбец 6 — Средняя цена
        price_cell = ws.cell(row=row_index, column=6)
        price_cell.alignment = number_alignment
        price_cell.number_format = "#,##0"

        # Столбец 7 — Занятость (%)
        occupancy_cell = ws.cell(row=row_index, column=7)
        occupancy_cell.alignment = number_alignment
        occupancy_cell.number_format = "0.0%"

        # Столбец 14 — Ссылка как гиперссылка
        link_cell = ws.cell(row=row_index, column=14)
        url = listing.full_url
        if url:
            link_cell.hyperlink = url
            link_cell.font = link_font

        # Чередующийся цвет строк
        if row_index % 2 == 0:
            even_fill = PatternFill(
                start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
            )
            for col_index in range(1, len(REPORT_COLUMNS) + 1):
                ws.cell(row=row_index, column=col_index).fill = even_fill


def apply_formatting(ws: Worksheet, data_rows: int) -> None:
    # Ширина столбцов
    for col_index, (_, width) in enumerate(REPORT_COLUMNS, start=1):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = width

    # Автофильтр
    last_col_letter = get_column_letter(len(REPORT_COLUMNS))
    last_row = data_rows + 1
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Фиксация шапки
    ws.freeze_panes = "A2"

    # Высота строки заголовка
    ws.row_dimensions[1].height = 30


def export_to_excel(listings: list[RawListing], output_path: str) -> str:
    """Формирует Excel-файл и возвращает абсолютный путь."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Аренда Avito"

    write_header(ws)
    write_data(ws, listings)
    apply_formatting(ws, len(listings))

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    absolute_path = str(out.resolve())
    return absolute_path


# ──────────────────────────────────────────────
# Точка входа
# ──────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Экспорт объявлений из SQLite в Excel (без парсинга)"
    )
    parser.add_argument(
        "--db",
        default="data/avito_products.db",
        help="Путь к файлу SQLite (по умолчанию: data/avito_products.db)",
    )
    parser.add_argument(
        "--out",
        default="data/avito_report_from_db.xlsx",
        help="Путь к выходному Excel-файлу (по умолчанию: data/avito_report_from_db.xlsx)",
    )
    args = parser.parse_args()

    print(f"БД:     {args.db}")
    print(f"Выход:  {args.out}")
    print("─" * 50)

    # 1. Читаем все объявления из БД
    listings = load_all_listings(args.db)

    if not listings:
        print("В базе данных нет объявлений. Экспорт пропущен.")
        sys.exit(0)

    # 2. Выводим краткую статистику
    categories: dict[str, int] = {}
    for listing in listings:
        cat = listing.room_category.value
        categories[cat] = categories.get(cat, 0) + 1

    print(f"\nСтатистика по категориям:")
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")

    avg_prices = [l.average_price for l in listings if l.average_price > 0]
    if avg_prices:
        print(f"\nСредняя цена (по всем объявлениям): {sum(avg_prices) / len(avg_prices):,.0f} руб./сут.")

    avg_occupancy = [l.occupancy_rate for l in listings if l.calendar_60_days]
    if avg_occupancy:
        print(f"Средняя занятость: {sum(avg_occupancy) / len(avg_occupancy) * 100:.1f}%")

    print("─" * 50)

    # 3. Экспортируем в Excel
    result_path = export_to_excel(listings, args.out)
    print(f"\nГотово! Excel-файл сохранён: {result_path}")
    print(f"Объявлений в файле: {len(listings)}")


if __name__ == "__main__":
    main()
