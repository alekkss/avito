"""Сервис экспорта данных в Excel.

Формирует Excel-файл с одним листом, содержащим все
объявления краткосрочной аренды. Таблица оптимизирована для
фильтрации и сортировки в Excel: автофильтры,
фиксированная шапка, автоширина столбцов.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config import ExportSettings, get_logger
from src.models import RawListing
from src.repositories.base import BaseListingRepository

logger = get_logger("export_service")

# Определение столбцов отчёта: (заголовок, ширина в символах)
REPORT_COLUMNS: list[tuple[str, int]] = [
    ("ID объявления", 18),
    ("Название", 45),
    ("Категория жилья", 16),
    ("Широта", 12),
    ("Долгота", 12),
    ("Средняя цена (руб./сут.)", 22),
    ("Занятость (%)", 14),
    ("Мин. срок (сут.)", 16),
    ("Мгновенное бронирование", 22),
    ("Рейтинг хоста", 14),
    ("Последнее обновление хостом", 26),
    ("Цены 60 дней", 80),       # увеличено: 60 значений через ";"
    ("Календарь 60 дней", 40),
    ("Ссылка", 55),
    ("Дата снимка", 20),
]


class ExportService:
    """Сервис для экспорта объявлений аренды в Excel-файл.

    Читает объявления из репозитория и формирует
    отформатированный Excel-файл с автофильтрами и стилями
    для удобной работы в Excel.

    Attributes:
        _repository: Репозиторий для чтения объявлений.
        _settings: Настройки экспорта (путь к файлу).
    """

    def __init__(
        self,
        repository: BaseListingRepository,
        settings: ExportSettings,
    ) -> None:
        """Инициализирует сервис экспорта.

        Args:
            repository: Репозиторий с объявлениями аренды.
            settings: Настройки экспорта (путь к выходному файлу).
        """
        self._repository = repository
        self._settings = settings

    def export(self) -> str:
        """Экспортирует все объявления аренды в Excel-файл.

        Основной публичный метод. Читает данные из репозитория,
        создаёт Excel-файл с форматированной таблицей и сохраняет
        по указанному пути.

        Returns:
            Абсолютный путь к созданному Excel-файлу.
            Пустая строка если нет данных для экспорта.
        """
        listings = self._repository.get_all_listings()

        if not listings:
            logger.warning("no_listings_to_export")
            return ""

        logger.info(
            "export_started",
            listings_count=len(listings),
            export_path=self._settings.export_path,
        )

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Аренда Avito"

        self._write_header(ws)
        self._write_data(ws, listings)
        self._apply_formatting(ws, len(listings))

        output_path = self._save_workbook(wb)

        categories = {listing.room_category.value for listing in listings}
        logger.info(
            "export_completed",
            listings_count=len(listings),
            unique_categories=len(categories),
            export_path=output_path,
        )

        return output_path

    def _write_header(self, ws: Worksheet) -> None:
        """Записывает строку заголовков в первую строку листа.

        Args:
            ws: Рабочий лист Excel.
        """
        header_font = Font(
            name="Calibri",
            bold=True,
            size=11,
            color="FFFFFF",
        )
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_index, (title, _) in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_index, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _write_data(
        self,
        ws: Worksheet,
        listings: list[RawListing],
    ) -> None:
        """Записывает данные объявлений в лист начиная со второй строки.

        Args:
            ws: Рабочий лист Excel.
            listings: Список объявлений аренды.
        """
        data_alignment = Alignment(
            vertical="top",
            wrap_text=False,
        )
        number_alignment = Alignment(
            horizontal="right",
            vertical="top",
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        link_font = Font(
            name="Calibri",
            size=10,
            color="0563C1",
            underline="single",
        )
        data_font = Font(
            name="Calibri",
            size=10,
        )

        for row_index, listing in enumerate(listings, start=2):
            row_data = self._listing_to_row(listing)

            for col_index, value in enumerate(row_data, start=1):
                cell = ws.cell(
                    row=row_index,
                    column=col_index,
                    value=value,
                )
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

            # Форматирование числовых столбцов
            # Столбец 6 — Средняя цена
            price_cell = ws.cell(row=row_index, column=6)
            price_cell.alignment = number_alignment
            price_cell.number_format = "#,##0"

            # Столбец 7 — Занятость (%)
            occupancy_cell = ws.cell(row=row_index, column=7)
            occupancy_cell.alignment = number_alignment
            occupancy_cell.number_format = "0.0%"

            # Столбец 14 — Ссылка как гиперссылка
            link_cell = ws.cell(row=row_index, column=14)
            url = listing.full_url
            if url:
                link_cell.hyperlink = url
                link_cell.font = link_font

            # Чередующийся цвет строк для читаемости
            if row_index % 2 == 0:
                even_fill = PatternFill(
                    start_color="D9E2F3",
                    end_color="D9E2F3",
                    fill_type="solid",
                )
                for col_index in range(1, len(REPORT_COLUMNS) + 1):
                    ws.cell(row=row_index, column=col_index).fill = even_fill

    def _listing_to_row(
        self, listing: RawListing
    ) -> list[str | int | float]:
        """Преобразует RawListing в список значений строки.

        Порядок значений соответствует порядку столбцов
        в REPORT_COLUMNS.

        Args:
            listing: Объявление аренды.

        Returns:
            Список значений для одной строки Excel.
        """
        snapshot_date = listing.snapshot_timestamp.strftime(
            "%Y-%m-%d %H:%M"
        )

        last_update = ""
        if listing.last_host_update is not None:
            last_update = listing.last_host_update.strftime(
                "%Y-%m-%d %H:%M"
            )

        prices_str = self._format_array_semicolon(listing.price_60_days)
        calendar_str = self._format_calendar_short(listing.calendar_60_days)

        return [
            listing.external_id,
            listing.title,
            listing.room_category.value,
            listing.latitude,
            listing.longitude,
            round(listing.average_price),
            listing.occupancy_rate,
            listing.min_stay,
            "Да" if listing.is_instant_book else "Нет",
            listing.host_rating,
            last_update,
            prices_str,
            calendar_str,
            listing.full_url,
            snapshot_date,
        ]

    def _format_array_semicolon(self, values: list[int]) -> str:
        """Форматирует массив цен в строку через точку с запятой.

        Все 60 значений выводятся в одну строку через «;».
        Занятые дни имеют значение 0, свободные — реальную цену.
        Пример: «0;0;0;5500;5500;4800;0;5500;...»

        Args:
            values: Массив целых чисел (цены на 60 дней).

        Returns:
            Строка всех значений через «;» или «—» если пусто.
        """
        if not values:
            return "—"
        return ";".join(str(v) for v in values)

    def _format_calendar_short(self, values: list[int]) -> str:
        """Форматирует массив занятости в компактную строку.

        Показывает первые 14 дней как последовательность 0/1.
        Например: «00011100001110... (60)»

        Args:
            values: Массив 0/1 (занятость).

        Returns:
            Компактная строковая запись календаря.
        """
        if not values:
            return "—"

        preview_count = 14
        preview = "".join(str(v) for v in values[:preview_count])

        if len(values) > preview_count:
            preview += f"... ({len(values)})"

        return preview

    def _apply_formatting(
        self, ws: Worksheet, data_rows: int
    ) -> None:
        """Применяет финальное форматирование к листу.

        Устанавливает ширину столбцов, автофильтры и фиксирует
        первую строку (шапку) для удобной прокрутки.

        Args:
            ws: Рабочий лист Excel.
            data_rows: Количество строк данных (без заголовка).
        """
        # Ширина столбцов
        for col_index, (_, width) in enumerate(REPORT_COLUMNS, start=1):
            col_letter = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = width

        # Автофильтр на все столбцы
        last_col_letter = get_column_letter(len(REPORT_COLUMNS))
        last_row = data_rows + 1
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        # Фиксация первой строки (заголовок)
        ws.freeze_panes = "A2"

        # Высота строки заголовка
        ws.row_dimensions[1].height = 30

        logger.debug(
            "formatting_applied",
            columns=len(REPORT_COLUMNS),
            data_rows=data_rows,
        )

    def _save_workbook(self, wb: Workbook) -> str:
        """Сохраняет Excel-файл на диск.

        Создаёт директорию для файла если она не существует.

        Args:
            wb: Объект Workbook для сохранения.

        Returns:
            Абсолютный путь к сохранённому файлу.
        """
        output_path = Path(self._settings.export_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(str(output_path))
        absolute_path = str(output_path.resolve())

        logger.info(
            "workbook_saved",
            path=absolute_path,
        )

        return absolute_path